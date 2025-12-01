import json
from openpyxl import load_workbook
import math
import os

# Excel dosyasını yükle - GitHub için relative path
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
wb = load_workbook(os.path.join(BASE_DIR, 'nota_cikarim_analizi_ilk.xlsx'))
ws = wb.active

# "Real Notes Count" sütununun indeksini bul
real_notes_col = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "Real Notes Count":
        real_notes_col = col
        break

# Model isimlerini ve sütun indekslerini al
model_columns = {}
for col in range(real_notes_col + 1, ws.max_column + 1):
    model_name = ws.cell(row=1, column=col).value
    if model_name:
        model_columns[model_name] = col

print("🔍 Gelişmiş Model Doğruluk Analizi Başlıyor...")
print("=" * 90)
print("📊 Hesaplanan Metrikler:")
metrics_info = [
    "1. Exact Match Accuracy: Tam eşleşme oranı",
    "2. Mean Absolute Error (MAE): Ortalama mutlak hata",
    "3. Root Mean Square Error (RMSE): Kök ortalama kare hata",
    "4. Weighted Accuracy: Hata mesafesine göre ağırlıklı skor",
    "🎯 YENİ KURAL: 0 nota çıkarımı = 0 puan (otomatik)"
]
for info in metrics_info:
    print(f"  {info}")
print("=" * 90)
print()

# Her model için detaylı analiz
results = {}

for model_name, col_idx in model_columns.items():
    exact_matches = 0
    total_count = 0
    absolute_errors = []
    squared_errors = []
    error_distribution = {0: 0, 1: 0, 2: 0, 3: 0, "4+": 0}
    zero_note_count = 0
    weighted_score = 0
    
    for row in range(2, ws.max_row + 1):
        real_value = ws.cell(row=row, column=real_notes_col).value or 0
        model_value = ws.cell(row=row, column=col_idx).value or 0
        
        total_count += 1
        
        if model_value == real_value:
            exact_matches += 1
        
        if model_value == 0:
            zero_note_count += 1
            weighted_score += 0
            error = real_value
            absolute_errors.append(error)
            squared_errors.append(error ** 2)
            error_distribution[error if error <= 3 else "4+"] += 1
        else:
            error = abs(real_value - model_value)
            absolute_errors.append(error)
            squared_errors.append(error ** 2)
            error_distribution[error if error <= 3 else "4+"] += 1
            
            score_map = {0: 100, 1: 75, 2: 50, 3: 25}
            weighted_score += score_map.get(error, 0)
    
    exact_accuracy = (exact_matches / total_count * 100) if total_count > 0 else 0
    mae = sum(absolute_errors) / len(absolute_errors) if absolute_errors else 0
    rmse = math.sqrt(sum(squared_errors) / len(squared_errors)) if squared_errors else 0
    weighted_accuracy = (weighted_score / (total_count * 100)) * 100
    
    results[model_name] = {
        'exact_matches': exact_matches,
        'total_count': total_count,
        'exact_accuracy': exact_accuracy,
        'mae': mae,
        'rmse': rmse,
        'weighted_accuracy': weighted_accuracy,
        'error_distribution': error_distribution,
        'zero_note_count': zero_note_count,
        'absolute_errors': absolute_errors
    }
    
    print(f"✓ {model_name}")
    print(f"  Exact Match: {exact_matches}/200 ({exact_accuracy:.2f}%)")
    print(f"  MAE: {mae:.3f}")
    print(f"  RMSE: {rmse:.3f}")
    print(f"  Weighted Score: {weighted_accuracy:.2f}%")
    print(f"  0 Nota Çıkarım: {zero_note_count} kez")
    print(f"  Hata Dağılımı: 0={error_distribution[0]}, 1={error_distribution[1]}, "
          f"2={error_distribution[2]}, 3={error_distribution[3]}, 4+={error_distribution['4+']}")
    print()

print("=" * 90)
print()

# Detaylı rapor oluştur
report = []

def add_section(title, width=110):
    """Bölüm başlığı ekle"""
    report.append("=" * width)
    report.append(title)
    report.append("=" * width)
    report.append("")

def add_subsection(title, width=110):
    """Alt bölüm başlığı ekle"""
    report.append(title)
    report.append("-" * width)

def add_lines(lines):
    """Satırları toplu ekle"""
    report.extend(lines)
    report.append("")

# BAŞLIK
add_section("GELİŞMİŞ MODEL DOĞRULUK ANALİZİ - HATA MESAFESİ DAHİL\n🎯 YENİ KURAL: 0 NOTA ÇIKARIMI = 0 PUAN")

# METRİK AÇIKLAMALARI
add_subsection("📊 METRİK AÇIKLAMALARI")
metric_descriptions = [
    ("1️⃣  EXACT MATCH ACCURACY (Tam Eşleşme Doğruluğu)", [
        "    → Model çıkardığı nota sayısı = Gerçek nota sayısı",
        "    → İkili (binary) metrik: Ya doğru ya yanlış",
        "    → Yanlışın büyüklüğünü dikkate almaz"
    ]),
    ("2️⃣  MAE (Mean Absolute Error - Ortalama Mutlak Hata)", [
        "    → Ortalama hata miktarı",
        "    → Formül: Σ|gerçek - tahmin| / n",
        "    → Düşük = İyi, Yüksek = Kötü",
        "    → Örnek: MAE=0.5 demek ortalama 0.5 nota farkı var"
    ]),
    ("3️⃣  RMSE (Root Mean Square Error - Kök Ortalama Kare Hata)", [
        "    → Büyük hatalara daha fazla ceza veren metrik",
        "    → Formül: √(Σ(gerçek - tahmin)² / n)",
        "    → MAE'den yüksekse büyük hatalar var demektir"
    ]),
    ("4️⃣  WEIGHTED ACCURACY (Ağırlıklı Doğruluk Skoru) ⭐ YENİ KURAL", [
        "    → Hata mesafesine göre puanlama sistemi:",
        "       • Model 0 nota çıkardı        = 0 puan ❌ (YENİ!)",
        "       • 0 hata (tam doğru)          = 100 puan ✅",
        "       • 1 hata (1 nota fark)        = 75 puan",
        "       • 2 hata (2 nota fark)        = 50 puan",
        "       • 3 hata (3 nota fark)        = 25 puan",
        "       • 4+ hata (4+ nota fark)      = 0 puan ❌",
        "    → Yüksek = İyi performans"
    ])
]

for title, lines in metric_descriptions:
    add_lines([title] + lines)

add_lines([
    "⚠️  ÖNEMLİ: 0 nota çıkarımı artık otomatik olarak 0 puan alıyor!",
    "    Örnek: Gerçekte 1 nota var, model 0 çıkardı → 0 puan (75 değil)"
])

# SIRALAMALAR
add_section("SIRALAMALAR")

# Sıralama türleri
rankings = [
    ("🎯 EXACT MATCH ACCURACY SIRALAMA", 
     sorted(results.items(), key=lambda x: x[1]['exact_accuracy'], reverse=True),
     lambda d: f"{d['exact_matches']}/200{'':<5} {d['exact_accuracy']:.2f}%",
     f"{'Sıra':<6} {'Model':<30} {'Doğru/Toplam':<15} {'Exact Acc.':<15}"),
    
    ("📉 MAE (ORTALAMA HATA) SIRALAMA - Düşük İyi",
     sorted(results.items(), key=lambda x: x[1]['mae']),
     lambda d: f"{d['mae']:.3f}{'':<10} Ortalama {d['mae']:.3f} nota hata yapıyor",
     f"{'Sıra':<6} {'Model':<30} {'MAE':<15} {'Açıklama':<50}"),
    
    ("📉 RMSE SIRALAMA - Düşük İyi",
     sorted(results.items(), key=lambda x: x[1]['rmse']),
     lambda d: f"{d['rmse']:.3f}{'':<10} +{d['rmse'] - d['mae']:.3f}",
     f"{'Sıra':<6} {'Model':<30} {'RMSE':<15} {'MAE ile Fark':<20}"),
    
    ("⭐ WEIGHTED ACCURACY SIRALAMA - Yüksek İyi (0 NOTA = 0 PUAN)",
     sorted(results.items(), key=lambda x: x[1]['weighted_accuracy'], reverse=True),
     lambda d: f"{d['weighted_accuracy']:.2f}%{'':<13} {d['zero_note_count']} kez",
     f"{'Sıra':<6} {'Model':<30} {'Weighted Score':<20} {'0 Nota Sayısı':<20}")
]

for title, sorted_data, format_func, header in rankings:
    add_subsection(title)
    report.append(header)
    add_subsection("")
    for idx, (model, data) in enumerate(sorted_data, 1):
        symbol = ["🥇", "🥈", "🥉"][idx-1] if idx <= 3 else f"{idx}."
        report.append(f"{symbol:<6} {model:<30} {format_func(data)}")
    report.append("")

# DETAYLI KARŞILAŞTIRMA TABLOSU
add_section("DETAYLI KARŞILAŞTIRMA TABLOSU")
report.append(f"{'Model':<30} {'Exact Acc':<12} {'MAE':<10} {'RMSE':<10} {'Weighted':<12} {'0 Nota':<10} {'Hata: 0|1|2|3|4+':<25}")
add_subsection("")

sorted_weighted = sorted(results.items(), key=lambda x: x[1]['weighted_accuracy'], reverse=True)
for model, data in sorted_weighted:
    err_dist = data['error_distribution']
    err_str = f"{err_dist[0]}|{err_dist[1]}|{err_dist[2]}|{err_dist[3]}|{err_dist['4+']}"
    report.append(f"{model:<30} {data['exact_accuracy']:.2f}%{'':<5} {data['mae']:.3f}{'':<5} "
                 f"{data['rmse']:.3f}{'':<5} {data['weighted_accuracy']:.2f}%{'':<5} {data['zero_note_count']:<10} {err_str}")
report.append("")

# ÖRNEK SENARYOLAR
add_section("🔍 ÖRNEK SENARYOLAR - Neden 0 Nota = 0 Puan Kuralı Önemli?")

scenarios = [
    ("Gerçek: 1 nota olmalı", [
        ("❌ ESKİ SISTEM (Yanlış):", [
            "  Senaryo: Model → 0 nota çıkardı",
            "  • Hata: |1-0| = 1",
            "  • Weighted Score: 75 puan",
            "  • Sorun: Hiç nota çıkaramayan model 75 puan alıyor!"
        ]),
        ("✅ YENİ SISTEM (Doğru):", [
            "  Senaryo: Model → 0 nota çıkardı",
            "  • Kontrol: model_value == 0?",
            "  • Weighted Score: 0 puan",
            "  • Sonuç: Hiç nota çıkaramayan model 0 puan alır!"
        ])
    ]),
    ("Gerçek: 4 nota olmalı", [
        ("Senaryo 1: Model A → 3 nota çıkardı", [
            "  • Hata: |4-3| = 1",
            "  • Weighted Score: 75 puan",
            "  • Yorum: İyi performans, sadece 1 nota eksik"
        ]),
        ("Senaryo 2: Model B → 0 nota çıkardı", [
            "  • 0 Nota Kontrolü: Evet → 0 puan",
            "  • Weighted Score: 0 puan",
            "  • Yorum: Hiç nota çıkaramadı, başarısız"
        ]),
        ("", ["❗ Model A (3 nota) = 75 puan, Model B (0 nota) = 0 puan → ADIL!"])
    ])
]

for scenario_title, scenario_parts in scenarios:
    report.append(scenario_title)
    report.append("")
    for part_title, part_lines in scenario_parts:
        if part_title:
            report.append(part_title)
        report.extend(part_lines)
        report.append("")
    report.append("-" * 110)
    report.append("")

# EN İYİ VS EN KÖTÜ
add_section("📊 EN İYİ vs EN KÖTÜ MODEL KARŞILAŞTIRMASI")

best_weighted = sorted_weighted[0]
worst_weighted = sorted_weighted[-1]

comparisons = [
    ("🏆 EN İYİ MODEL", best_weighted),
    ("❌ EN KÖTÜ MODEL", worst_weighted)
]

for title, (model_name, data) in comparisons:
    report.append(f"{title}: {model_name}")
    metrics = [
        f"   Exact Match Accuracy: {data['exact_accuracy']:.2f}%",
        f"   MAE (Ortalama Hata): {data['mae']:.3f} nota",
        f"   RMSE: {data['rmse']:.3f}",
        f"   Weighted Accuracy: {data['weighted_accuracy']:.2f}%",
        f"   0 Nota Çıkarım: {data['zero_note_count']} kez",
        f"   Hata Dağılımı: {data['error_distribution']}"
    ]
    report.extend(metrics)
    report.append("")

report.append("📈 PERFORMANS FARKI:")
performance_diff = [
    f"   Exact Match Farkı: {best_weighted[1]['exact_accuracy'] - worst_weighted[1]['exact_accuracy']:.2f}%",
    f"   MAE Farkı: {worst_weighted[1]['mae'] - best_weighted[1]['mae']:.3f} nota",
    f"   Weighted Accuracy Farkı: {best_weighted[1]['weighted_accuracy'] - worst_weighted[1]['weighted_accuracy']:.2f}%",
    f"   0 Nota Farkı: {worst_weighted[1]['zero_note_count'] - best_weighted[1]['zero_note_count']} kez"
]
report.extend(performance_diff)
report.append("")

# SONUÇ VE ÖNERİLER
add_section("🎯 SONUÇ VE ÖNERİLER")

recommendations = [
    ("1️⃣  HANGİ METRİK KULLANILMALI?", [
        "   • EXACT MATCH: Sadece 'doğru/yanlış' bilgisi istiyorsan",
        "   • MAE: Ortalama hata miktarını görmek istiyorsan",
        "   • RMSE: Büyük hataların etkisini görmek istiyorsan",
        "   • WEIGHTED ACCURACY: Hem doğruluğu hem hata mesafesini dengeli değerlendirmek istiyorsan",
        "",
        "   ✅ ÖNERİ: WEIGHTED ACCURACY en kapsamlı metriktir!",
        "   ✅ YENİ: 0 nota çıkarımı artık adil şekilde cezalandırılıyor"
    ]),
    ("2️⃣  ÜRETİM İÇİN MODEL SEÇİMİ:", [
        f"   🏆 Weighted Accuracy'ye göre: {best_weighted[0]}",
        "      → En az hata yapan model",
        "      → Yaptığı hatalar da daha küçük",
        f"      → 0 nota çıkarım: Sadece {best_weighted[1]['zero_note_count']} kez"
    ]),
    ("3️⃣  0 NOTA KURALI ETKİSİ:", [
        "   Bu kural özellikle şu durumlarda önemli:",
        "   • Gerçekte az nota var (1-2 nota)",
        "   • Model hiç nota çıkaramıyor",
        "   • Eski sistemde: 1 hata = 75 puan alırdı (yanıltıcı)",
        "   • Yeni sistemde: 0 nota = 0 puan (adil)",
        "",
        f"   🔍 Analiz: En kötü model {worst_weighted[1]['zero_note_count']} kez 0 nota çıkardı",
        f"            En iyi model {best_weighted[1]['zero_note_count']} kez 0 nota çıkardı"
    ])
]

for title, lines in recommendations:
    report.append(title)
    report.append("")
    report.extend(lines)
    report.append("")

report.append("=" * 110)
report.append("Rapor Oluşturulma Tarihi: 23 Kasım 2024")
report.append("=" * 110)

# Dosyaya kaydet
output_txt = os.path.join(BASE_DIR, 'gelismis_dogruluk_analizi_yeni_kural.txt')
report_text = "\n".join(report)
with open(output_txt, 'w', encoding='utf-8') as f:
    f.write(report_text)

# JSON formatında kaydet
json_results = {
    'summary': {
        'total_reviews': 200,
        'total_models': len(results),
        'new_rule': 'Zero note extraction = 0 points',
        'metrics': {
            'exact_match': 'Binary accuracy (correct/incorrect)',
            'mae': 'Mean Absolute Error (lower is better)',
            'rmse': 'Root Mean Square Error (lower is better)',
            'weighted_accuracy': 'Weighted score with zero-note penalty (higher is better)'
        }
    },
    'models': {
        model_name: {
            'exact_match_accuracy': round(data['exact_accuracy'], 2),
            'exact_matches': data['exact_matches'],
            'mae': round(data['mae'], 3),
            'rmse': round(data['rmse'], 3),
            'weighted_accuracy': round(data['weighted_accuracy'], 2),
            'zero_note_count': data['zero_note_count'],
            'error_distribution': {
                '0_errors': data['error_distribution'][0],
                '1_error': data['error_distribution'][1],
                '2_errors': data['error_distribution'][2],
                '3_errors': data['error_distribution'][3],
                '4plus_errors': data['error_distribution']['4+']
            }
        }
        for model_name, data in results.items()
    }
}

output_json = os.path.join(BASE_DIR, 'gelismis_dogruluk_analizi_yeni_kural.json')
with open(output_json, 'w', encoding='utf-8') as f:
    json.dump(json_results, f, ensure_ascii=False, indent=2)

print("\n✅ Gelişmiş analiz raporları oluşturuldu!")
print(f"📄 Detaylı Rapor: {output_txt}")
print(f"📊 JSON Rapor: {output_json}")
print("\n" + "=" * 90)
print("WEIGHTED ACCURACY SIRALAMA (0 NOTA = 0 PUAN KURALI):")
print("=" * 90)
for idx, (model, data) in enumerate(sorted_weighted, 1):
    symbol = ["🥇", "🥈", "🥉"][idx-1] if idx <= 3 else f"{idx}."
    print(f"{symbol} {model:<30} → {data['weighted_accuracy']:.2f}% (MAE: {data['mae']:.3f}, 0 nota: {data['zero_note_count']})")
print("=" * 90)